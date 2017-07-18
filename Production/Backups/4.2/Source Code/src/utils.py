#! python3
# -*- coding: utf-8 -*-

"""utils: A collection of methods for the program.

__author__ = "Katie Nelson"
__email__  = [ "KatherineE.Nelson@scotiabank.com", "ketnscripts@gmail.com" ]

"""
import os
import xlrd
import openpyxl
from datetime import datetime

def convert_xls(xls_path, testing=False):
    """Converts .xls files to .xlsx.

    Args:
        xls_path (string): the full path string to the .xls file.

    Returns:
        An open Workdbook() openpyxl class populated from the .xls file.

    Raises:
        ValueError: if the path passed either does not exist, is not
            an absolute path, or does not point to a .xls file
    """
    if not os.path.isabs(xls_path):
        raise ValueError('Non-absolute path passed to excel_utils convert method.')
    if not os.path.lexists(xls_path):
        raise ValueError('Passed path does not exist.')
    if not xls_path.endswith('.xls'):
        raise ValueError('Passed path is not a .xls file.')
    if testing:
        return 'all good'

    xls_file = xlrd.open_workbook(xls_path)
    new_file = openpyxl.Workbook()

    for i in range(0, xls_file.nsheets):
        sheet = xls_file.sheet_by_index(i)
        # active sheet if its the first, otherwise make another sheet
        active = new_file.active if i==0 else new_file.create_sheet()
        active.title = sheet.name
        for row in range(0, sheet.nrows):
            for col in range(0, sheet.ncols):
                value = sheet.cell_value(row, col)

                #if cell is a date object, save it as such
                if sheet.cell_type(row, col) == 3:
                    date_tuple = (xlrd.xldate.xldate_as_tuple(value, 0))
                    value = datetime(*(date_tuple))

                active.cell(row=row+1, column=col+1).value = value

    return new_file
