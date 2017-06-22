#! python3
# -*- coding: utf-8 -*-
"""Excel_Control: Core excel functionality

__author__ = "Katie Nelson"
__email__  = [ "KatherineE.Nelson@scotiabank.com", "ketnscripts@gmail.com" ]

"""
import os, glob
import shutil
import src
import src.utils as utils
import openpyxl
from datetime import datetime
import math

class Excel_Control(object):

    def __init__(self):
        """Constructor for the excel_control object

        NOTE: input automates to the newest created/modifies .xls(x) file
        """
        self.output_path = os.path.join(src.args.output_dir, src.args.output_name) + '.xlsx'
        self.input_path  = max(glob.iglob(os.path.join(src.args.input_dir, '*.xls*')), key=os.path.getctime)
        self.template    = src.args.template
        self.input_file  = self.__get_input_file()
        self.output_file = self.__get_output_file()
        self.backup_dir  = src.args.backup_dir if src.args.backup else None

    def __get_input_file(self):
        """Private method: get_input_file
        """
        if self.input_path.endswith('s'):
            try:
                ret = utils.convert_xls(self.input_path)
                self.input_path += 'x'
                return ret
            except ValueError as err:
                sys.exit('Your input path is incorrectly formatted. Traceback:', err)

        else:
            return openpyxl.load_workbook(self.input_path)

    def __get_output_file(self):
        """Private method: get_output_file
        """
        shutil.copy(self.template, self.output_path)
        return openpyxl.load_workbook(self.output_path)

    def populate(self):
        """Populate: Populates the excel template with data

        this is VERY specific to the template bundled in with the program

        changing templates will involve manipulating source code directly.
        """
        in_sheet = self.input_file.active
        out_sheet = self.output_file.active

        for i in range(2, 90):
            country = in_sheet.cell(row=i, column=2).value
            country = country.replace(' ', '_').replace('.', '')

            task = in_sheet.cell(row=i, column=3).value
            task = task.replace(' ', '_').replace('-', '_')

            tmp = in_sheet.cell(row=i, column=4).value
            start_time = tmp if type(tmp) is datetime else ''

            tmp = in_sheet.cell(row=i, column=5).value
            end_time = tmp if type(tmp) is datetime else ''

            tmp = float(in_sheet.cell(row=i, column=6).value)


            elapsed = math.ceil(((end_time - start_time).total_seconds()) / 60) if \
                 start_time and end_time else ''

            # gets the correct row from our config package
            index = getattr(src.index, task)
            r = index[country]

            start_time = start_time.strftime('%H:%M') if type(start_time) is datetime else ''
            end_time = end_time.strftime('%H:%M') if type(end_time) is datetime else ''
            out_sheet.cell(row=r, column=3).value = start_time
            out_sheet.cell(row=r, column=4).value = end_time
            out_sheet.cell(row=r, column=5).value = elapsed

    def backup(self):
        """Creates a backup of files in the specified backup directory. """
        files = os.listdir(os.path.dirname(self.input_path))
        for item in files:
            shutil.copy(os.path.join(os.path.dirname(self.input_path), item), self.backup_dir)

    def save(self):
        """Cleans and saves data"""
        if (self.backup_dir):
            self.backup()

        files = os.listdir(os.path.dirname(self.input_path))
        for item in files:
            os.unlink(os.path.join(os.path.dirname(self.input_path), item))

        self.output_file.save(self.output_path)
