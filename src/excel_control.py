#! python3
# -*- coding: utf-8 -*-
"""Excel_Control: Core excel functionality

The meat of the program. This is very specific to the template.xlsx file and the
current (June 22, 2017) version of GEMS' output (our input) file.

"""
import os
import shutil
import math
import openpyxl
from glob import iglob
from datetime import datetime

import src
import src.config.index as excel_rows
import src.utils as utils


'''
Entry point to this file is below the Excel_Control class
'''

class Excel_Control(object):

    def __init__(self):
        """Constructor for the excel_control object

        NOTE: input automates to the newest created/modified .xls(x) file
        """

        # Catches an exception thrown from max() when it has no input
        try:
            self.input_path = max(iglob(os.path.join(src.args.input_dir, '*.xls*')),
                                  key=os.path.getctime)
        except:
            input('Failed to run: No files in input directory. <Press enter to exit>')
            sys.exit()

        self.output_path = os.path.join(src.args.output_dir,
                                        (src.args.output_name + '.xlsx'))
        self.template    = src.args.template
        self.input_file  = self.__get_input_file()
        self.output_file = self.__get_output_file()
        self.backup_dir  = src.args.backup_dir if src.args.backup else None
        self.index       = excel_rows.Index()

    def __get_input_file(self):
        """Private method: get_input_file
        """
        # if the input file is .xls, we convert it to .xlsx
        if self.input_path.endswith('s'):
            try:
                ret = utils.convert_xls(self.input_path)
                self.input_path += 'x'
                return ret
            except ValueError as err:
                sys.exit('Your input path is incorrectly formatted:', err)

        else:
            # returns an open excel workbook
            return openpyxl.load_workbook(self.input_path)

    def __get_output_file(self):
        """Private method: get_output_file
        """
        # Copies the template to the output path
        shutil.copy(self.template, self.output_path)
        #returns an open excel workbook
        return openpyxl.load_workbook(self.output_path)

    def populate(self):
        """Populate: Populates the excel template with data

        This is VERY specific to the template bundled in with the program

        Changing templates will involve manipulating source code directly.
        """
        # We're really only working with the first ('active') excel sheet
        in_sheet = self.input_file.active
        out_sheet = self.output_file.active

        # this runs for each row in the input file, except the first,
        # which is the column titles

        for i in range(2, 90):
            # the country is column 2, and we're sanitizing it so all spaces and
            # periods are removed. This corresponds to an index in our dictionary
            # (./src/config/index.py)
            country = in_sheet.cell(row=i, column=2).value
            country = country.replace(' ', '_').replace('.', '')

            # same idea for the task
            task = in_sheet.cell(row=i, column=3).value
            task = task.replace(' ', '_').replace('-', '_')

            # we're storing the time to make the next few lines more readable
            tmp = in_sheet.cell(row=i, column=4).value
            # two possible cases for the time--a datetime object or nothing
            # if it's nothing, start_time will be an empty string
            start_time = tmp if type(tmp) is datetime else ''

            tmp = in_sheet.cell(row=i, column=5).value
            end_time = tmp if type(tmp) is datetime else ''

            # if we have both a start and end time,
            # calculate the difference in minutes. We could also grab that from
            # the input file, but it often seems to be inaccurate, so we'll just
            # do it ourselves. If there is an issue calculating the time, advise
            # the operator to check the start/end times.
            if start_time and end_time:
                elapsed = (end_time - start_time).total_seconds()
                elapsed = math.ceil(elapsed / 60)
                if (elapsed > 1400) or (elapsed < 0):
                    elapsed = 'CHECK START/END TIME'
            else:
                elapsed = ''


            # Now we'll convert our datetime objects to formatted strings
            start_time = start_time.strftime('%H:%M') \
                         if type(start_time) is datetime else ''

            end_time = end_time.strftime('%H:%M') \
                       if type(end_time) is datetime else ''

            # these next few lines may need reworking

            # this retrieves the appropriate dictionary of excel rows for the
            # specific task from our index object
            task_rows = getattr(self.index, task)

            # we then grab the row specific to the country we're working on
            r = task_rows[country]

            # now we know what row we need, and the columns never change, so
            # we can put our data in the right place in our output file:

            out_sheet.cell(row=r, column=3).value = start_time
            out_sheet.cell(row=r, column=4).value = end_time
            out_sheet.cell(row=r, column=5).value = elapsed

            # and then we repeat for the next row

    # and that's all the formatting that needs to be done!

    def backup(self):
        """Creates a backup of files in the specified backup directory. """
        # get the input directory from it's path
        input_dir = os.path.dirname(self.input_path)

        # grab all the files in ./Runtime/Input..
        files = os.listdir(input_dir)
        for item in files:
            # ..and then copy them to the backup directory
            item_path = os.path.join(input_dir, item)
            shutil.copy(item_path, self.backup_dir)

    def save(self):
        """Cleans and saves data"""

        input_dir = os.path.dirname(self.input_path)
        # get all files in the input directory
        files = os.listdir(input_dir)
        for item in files:
            item_path = os.path.join(input_dir, item)
            # and delete them
            os.unlink(item_path)

        self.output_file.save(self.output_path)

        # if we're running in quick mode, don't open the excel file at the end
        if (src.args.quick):
            return

        # We're all done--open the file
        cmd = "start excel.exe \"" + self.output_path+"\""
        os.system(cmd)

# entry point to this file
def start():
    control = Excel_Control()
    control.populate()
    if (src.args.backup):
        control.backup()
    control.save()
